"""Excel report generator."""
from __future__ import annotations

import io
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import ReconcileResult


HEADER_FILL = PatternFill("solid", fgColor="1A365D")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="F1F5F9")
GREEN = PatternFill("solid", fgColor="C6F6D5")
YELLOW = PatternFill("solid", fgColor="FEFCBF")
RED = PatternFill("solid", fgColor="FED7D7")
THIN = Side(style="thin", color="CBD5E0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _derive_field(row: Dict[str, Any], field: str) -> Any:
    """Resolve a column value, including v3 rationale-derived fields.

    `Rationale` (top evidence line) and `Your note` (user_reason) live inside
    `row["rationale"]` rather than as top-level keys; this helper pulls them
    out so the table writer doesn't need to know the shape.
    """
    if field == "Rationale":
        rat = row.get("rationale") or {}
        ev = (rat.get("rationale") or [])
        if not ev:
            return None
        top = ev[0]
        return f"[{top.get('source')}] {top.get('evidence')}"
    if field == "Your note":
        rat = row.get("rationale") or {}
        return rat.get("user_reason") or ""
    if field == "Confidence":
        rat = row.get("rationale") or {}
        c = rat.get("confidence")
        return None if c is None else round(float(c), 2)
    return row.get(field)


def _write_table(ws, headers: List[str], rows: List[Dict[str, Any]], status_col: str | None = None):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER
    for i, row in enumerate(rows, 2):
        for j, h in enumerate(headers, 1):
            v = _derive_field(row, h)
            c = ws.cell(row=i, column=j, value=v)
            c.border = BORDER
            if i % 2 == 0:
                c.fill = ALT_FILL
            if status_col and h == status_col:
                s = (v or "").lower() if isinstance(v, str) else ""
                if s == "match":
                    c.fill = GREEN
                elif s in ("minor", "fee_offset"):
                    c.fill = YELLOW
                elif s == "major":
                    c.fill = RED
            # Wrap long rationale text so it stays readable
            if h in ("Rationale", "Your note"):
                c.alignment = Alignment(wrap_text=True, vertical="top")
    # auto width
    for j, h in enumerate(headers, 1):
        sample_vals = [str(_derive_field(r, h) or "") for r in rows[:200]]
        max_len = max([len(str(h))] + [len(v) for v in sample_vals])
        if h == "Rationale":
            ws.column_dimensions[get_column_letter(j)].width = 60
        elif h == "Your note":
            ws.column_dimensions[get_column_letter(j)].width = 30
        elif h == "Confidence":
            ws.column_dimensions[get_column_letter(j)].width = 12
        else:
            ws.column_dimensions[get_column_letter(j)].width = min(max(12, max_len + 2), 40)


def build_report(result: ReconcileResult) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    s = result.summary
    cfg = result.config
    summary_rows = [
        ("ReconOps AI — Reconciliation Report", ""),
        ("Job ID", result.job_id),
        ("Created", result.created_at),
        ("Type", cfg.recon_type),
        ("Source A", cfg.label_a),
        ("Source B", cfg.label_b),
        ("", ""),
        ("Total records — Source A", s.total_a),
        ("Total records — Source B", s.total_b),
        ("Matched", f"{s.matched} ({s.matched_pct}%)"),
        ("Fuzzy matches", s.fuzzy_matches),
        ("Unmatched (A only)", s.unmatched_a),
        ("Unmatched (B only)", s.unmatched_b),
        ("Amount discrepancies", s.discrepancies),
        ("Total discrepancy value ($)", s.total_discrepancy_value),
        ("Total amount — Source A ($)", s.total_amount_a),
        ("Total amount — Source B ($)", s.total_amount_b),
    ]
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    for i, (k, v) in enumerate(summary_rows, 1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=(i == 1 or k in {"Job ID", ""}))
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 40

    matched_headers = [
        "key", "match_type", "status", "Confidence", "fee_pattern",
        "amount_a", "amount_b", "diff_abs", "diff_pct",
        "date_a", "date_b", "delta_days",
        "Rationale", "Your note",
    ]
    ws2 = wb.create_sheet("Matched")
    _write_table(ws2, matched_headers, result.matched, status_col="status")

    if result.unmatched_a:
        ws3 = wb.create_sheet(f"Unmatched - {cfg.label_a[:20]}")
        headers = list(result.unmatched_a[0].keys())
        _write_table(ws3, headers, result.unmatched_a)

    if result.unmatched_b:
        ws4 = wb.create_sheet(f"Unmatched - {cfg.label_b[:20]}")
        headers = list(result.unmatched_b[0].keys())
        _write_table(ws4, headers, result.unmatched_b)

    ws5 = wb.create_sheet("Discrepancies")
    # Same column shape as Matched so rationale + note are present on the
    # tab the user spends the most time in.
    _write_table(ws5, matched_headers, result.discrepancies, status_col="status")

    ws6 = wb.create_sheet("Insights")
    ws6.cell(row=1, column=1, value="AI Insights").font = Font(bold=True, size=14)
    ws6.column_dimensions["A"].width = 110
    for i, line in enumerate(result.insights.split("\n"), 3):
        c = ws6.cell(row=i, column=1, value=line)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
